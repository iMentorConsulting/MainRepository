from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import and_
from database import get_db
from models import Booking, Unit, Customer
from schemas import BookingCreate, BookingUpdate, BookingResponse
from auth_utils import get_tenant
from typing import List, Optional
from datetime import date
from io import BytesIO

router = APIRouter(prefix="/bookings", tags=["bookings"])

CHANNEL_LABELS = {
    'booking': 'Booking.com', 'airbnb': 'Airbnb', 'direct': 'Απευθείας',
    'oga': 'ΟΓΑ', 'social_tourism': 'Κοιν.Τουρισμός', 'other': 'Άλλο',
}
CHANNEL_MAP_IMPORT = {
    'booking.com': 'booking', 'booking': 'booking', 'airbnb': 'airbnb',
    'απευθείας': 'direct', 'direct': 'direct', 'oga': 'oga', 'οga': 'oga',
    'κοιν.τουρισμός': 'social_tourism', 'social_tourism': 'social_tourism',
    'other': 'other', 'άλλο': 'other',
}
STATUS_LABELS = {
    'confirmed': 'Επιβεβαιωμένη', 'pending': 'Εκκρεμής',
    'cancelled': 'Ακυρωμένη', 'no_show': 'No-show',
}


def _check_overlap(db, unit_id, check_in, check_out, tenant, exclude_id=None):
    q = db.query(Booking).filter(
        Booking.unit_id == unit_id,
        Booking.tenant == tenant,
        Booking.status.in_(["confirmed", "pending"]),
        and_(Booking.check_in < check_out, Booking.check_out > check_in),
    )
    if exclude_id:
        q = q.filter(Booking.id != exclude_id)
    return q.first()


def _load(db, booking_id, tenant):
    return (
        db.query(Booking)
        .options(joinedload(Booking.unit), joinedload(Booking.customer))
        .filter(Booking.id == booking_id, Booking.tenant == tenant)
        .first()
    )


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    tenant: str = Depends(get_tenant),
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    unit_id: Optional[int] = None,
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    q = db.query(Booking).options(joinedload(Booking.unit), joinedload(Booking.customer)).filter(Booking.tenant == tenant)
    if from_date:
        q = q.filter(Booking.check_out >= from_date)
    if to_date:
        q = q.filter(Booking.check_in <= to_date)
    if unit_id:
        q = q.filter(Booking.unit_id == unit_id)
    bookings = q.order_by(Booking.check_in).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Κρατήσεις"

    headers = ["ID", "Μονάδα", "Πελάτης", "Κανάλι", "Άφιξη", "Αναχώρηση",
               "Νύχτες", "Άτομα", "Σύνολο (€)", "Προμήθεια (€)", "Καθαρό (€)",
               "Κατάσταση", "Τιμολογήθηκε", "Σημειώσεις"]
    ws.append(headers)

    blue = PatternFill("solid", fgColor="3B82F6")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = blue
        cell.alignment = Alignment(horizontal="center")

    for b in bookings:
        nights = (b.check_out - b.check_in).days
        name = f"{b.customer.first_name} {b.customer.last_name}".strip()
        ws.append([
            b.id, b.unit.name, name,
            CHANNEL_LABELS.get(b.channel, b.channel),
            b.check_in.strftime("%d/%m/%Y"), b.check_out.strftime("%d/%m/%Y"),
            nights, b.guests, b.total_price, b.commission,
            round(b.total_price - b.commission, 2),
            STATUS_LABELS.get(b.status, b.status),
            "Ναι" if b.is_billed else "Όχι", b.notes or "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=kratiseis.xlsx"})


@router.get("/template/excel")
def download_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    headers = ["Μονάδα", "Πελάτης", "Κανάλι", "Άφιξη (ΗΗ/ΜΜ/ΕΕΕΕ)",
               "Αναχώρηση (ΗΗ/ΜΜ/ΕΕΕΕ)", "Άτομα", "Σύνολο (€)",
               "Προμήθεια (€)", "Τιμολογήθηκε (Ναι/Όχι)", "Σημειώσεις"]
    ws.append(headers)
    blue = PatternFill("solid", fgColor="3B82F6")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = blue
    ws.append(["Διαμέρισμα 1", "Νίκος Παπαδόπουλος", "booking", "01/07/2026", "08/07/2026", 2, 700, 105, "Όχι", ""])
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=template_kratiseis.xlsx"})


@router.post("/import/excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db), tenant: str = Depends(get_tenant)):
    import openpyxl
    from datetime import datetime as dt

    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    created = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue
        try:
            (unit_name, customer_name, channel, check_in_str, check_out_str,
             guests, total_price, commission, is_billed_str, notes) = (list(row) + [None] * 10)[:10]

            unit = db.query(Unit).filter(Unit.name.ilike(str(unit_name).strip()), Unit.tenant == tenant).first()
            if not unit:
                errors.append(f"Γραμμή {row_idx}: Μονάδα '{unit_name}' δεν βρέθηκε")
                continue

            parts = str(customer_name or "Άγνωστος").strip().split(" ", 1)
            fname, lname = parts[0], (parts[1] if len(parts) > 1 else "")
            customer = db.query(Customer).filter(
                Customer.first_name.ilike(fname), Customer.last_name.ilike(lname), Customer.tenant == tenant
            ).first()
            if not customer:
                customer = Customer(first_name=fname, last_name=lname, tenant=tenant)
                db.add(customer)
                db.flush()

            if isinstance(check_in_str, str):
                check_in = dt.strptime(check_in_str.strip(), "%d/%m/%Y").date()
                check_out = dt.strptime(check_out_str.strip(), "%d/%m/%Y").date()
            else:
                check_in = check_in_str
                check_out = check_out_str

            ch = CHANNEL_MAP_IMPORT.get(str(channel or "").lower().strip(), "other")
            billed = str(is_billed_str or "").lower() in ["ναι", "yes", "true", "1"]

            booking = Booking(
                unit_id=unit.id, customer_id=customer.id, channel=ch,
                check_in=check_in, check_out=check_out,
                guests=int(guests or 1), total_price=float(total_price or 0),
                commission=float(commission or 0), status="confirmed",
                is_billed=billed, notes=str(notes or ""), tenant=tenant,
            )
            db.add(booking)
            created += 1
        except Exception as e:
            errors.append(f"Γραμμή {row_idx}: {str(e)}")

    db.commit()
    return {"created": created, "errors": errors}


@router.get("/", response_model=List[BookingResponse])
def get_bookings(
    db: Session = Depends(get_db),
    tenant: str = Depends(get_tenant),
    unit_id: Optional[int] = None,
    channel: Optional[str] = None,
    status: Optional[str] = None,
    is_billed: Optional[bool] = None,
    from_date: Optional[date] = None,
    to_date: Optional[date] = None,
    sort_by: str = "check_in",
    sort_dir: str = "desc",
    skip: int = 0,
    limit: int = 500,
):
    q = db.query(Booking).options(joinedload(Booking.unit), joinedload(Booking.customer)).filter(Booking.tenant == tenant)
    if unit_id:
        q = q.filter(Booking.unit_id == unit_id)
    if channel:
        q = q.filter(Booking.channel == channel)
    if status:
        q = q.filter(Booking.status == status)
    if is_billed is not None:
        q = q.filter(Booking.is_billed == is_billed)
    if from_date:
        q = q.filter(Booking.check_out >= from_date)
    if to_date:
        q = q.filter(Booking.check_in <= to_date)

    sort_col = {
        "check_in": Booking.check_in, "check_out": Booking.check_out,
        "total_price": Booking.total_price, "created_at": Booking.created_at,
    }.get(sort_by, Booking.check_in)
    q = q.order_by(sort_col.asc() if sort_dir == "asc" else sort_col.desc())
    return q.offset(skip).limit(limit).all()


@router.post("/", response_model=BookingResponse, status_code=201)
def create_booking(booking: BookingCreate, db: Session = Depends(get_db), tenant: str = Depends(get_tenant)):
    if booking.check_out <= booking.check_in:
        raise HTTPException(status_code=400, detail="Η αναχώρηση πρέπει να είναι μετά την άφιξη")
    overlap = _check_overlap(db, booking.unit_id, booking.check_in, booking.check_out, tenant)
    if overlap:
        raise HTTPException(status_code=409, detail=f"Σύγκρουση με κράτηση #{overlap.id} ({overlap.check_in} – {overlap.check_out})")
    obj = Booking(**booking.model_dump(), tenant=tenant)
    db.add(obj)
    db.flush()
    # Auto-create guest portal token
    from models import GuestToken
    gt = GuestToken(booking_id=obj.id, tenant=tenant)
    db.add(gt)
    db.commit()
    db.refresh(obj)
    return _load(db, obj.id, tenant)


@router.get("/{booking_id}", response_model=BookingResponse)
def get_booking(booking_id: int, db: Session = Depends(get_db), tenant: str = Depends(get_tenant)):
    obj = _load(db, booking_id, tenant)
    if not obj:
        raise HTTPException(status_code=404, detail="Κράτηση δεν βρέθηκε")
    return obj


@router.put("/{booking_id}", response_model=BookingResponse)
def update_booking(booking_id: int, data: BookingUpdate, db: Session = Depends(get_db), tenant: str = Depends(get_tenant)):
    obj = db.query(Booking).filter(Booking.id == booking_id, Booking.tenant == tenant).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Κράτηση δεν βρέθηκε")
    upd = data.model_dump(exclude_none=True)
    new_unit = upd.get("unit_id", obj.unit_id)
    new_in = upd.get("check_in", obj.check_in)
    new_out = upd.get("check_out", obj.check_out)
    if "check_in" in upd or "check_out" in upd or "unit_id" in upd:
        if new_out <= new_in:
            raise HTTPException(status_code=400, detail="Η αναχώρηση πρέπει να είναι μετά την άφιξη")
        overlap = _check_overlap(db, new_unit, new_in, new_out, tenant, exclude_id=booking_id)
        if overlap:
            raise HTTPException(status_code=409, detail=f"Σύγκρουση με κράτηση #{overlap.id}")
    for k, v in upd.items():
        setattr(obj, k, v)
    db.commit()
    return _load(db, booking_id, tenant)


@router.delete("/{booking_id}")
def delete_booking(booking_id: int, db: Session = Depends(get_db), tenant: str = Depends(get_tenant)):
    obj = db.query(Booking).filter(Booking.id == booking_id, Booking.tenant == tenant).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Κράτηση δεν βρέθηκε")
    db.delete(obj)
    db.commit()
    return {"message": "Η κράτηση διαγράφηκε"}
